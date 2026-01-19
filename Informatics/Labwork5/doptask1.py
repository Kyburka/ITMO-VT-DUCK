from openpyxl import load_workbook
import pandas as pd

wb = load_workbook("Informatics/Labwork5/lab5.xlsx", data_only=True)
ws = wb.active

data = []
column_letters = "ABCEGHIJKLMNOPQRSTUVWXY"

for row in ws.iter_rows(min_row=8, max_row=19):
    row_values = []
    for cell in row:
        if cell.column_letter not in column_letters:
            continue
        row_values.append(cell.value)
    data.append(row_values)

df = pd.DataFrame(data)
print(df)